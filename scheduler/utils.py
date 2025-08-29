"""
Utility functions for JW Attendant Scheduler
"""

import csv
import io
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def export_attendants_csv(attendants):
    """Export attendants to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendants_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Name', 'Email', 'Phone', 'Congregation', 'Serving As', 
        'Preferred Positions', 'Emergency Contact', 'Emergency Phone', 'Created Date'
    ])
    
    for attendant in attendants:
        writer.writerow([
            attendant.get_full_name(),
            attendant.email,
            attendant.phone or '',
            attendant.congregation or '',
            attendant.get_serving_as_display(),
            attendant.preferred_positions or '',
            attendant.emergency_contact or '',
            attendant.emergency_phone or '',
            attendant.created_at.strftime('%Y-%m-%d')
        ])
    
    return response


def export_attendants_excel(attendants):
    """Export attendants to Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendants"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Name', 'Email', 'Phone', 'Congregation', 'Serving As', 
        'Experience Level', 'Preferred Positions', 'Emergency Contact', 
        'Emergency Phone', 'Created Date'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, attendant in enumerate(attendants, 2):
        ws.cell(row=row, column=1, value=attendant.get_full_name())
        ws.cell(row=row, column=2, value=attendant.email)
        ws.cell(row=row, column=3, value=attendant.phone or '')
        ws.cell(row=row, column=4, value=attendant.congregation or '')
        ws.cell(row=row, column=5, value=attendant.get_serving_as_display())
        ws.cell(row=row, column=6, value=attendant.get_experience_level_display())
        ws.cell(row=row, column=7, value=attendant.preferred_positions or '')
        ws.cell(row=row, column=8, value=attendant.emergency_contact or '')
        ws.cell(row=row, column=9, value=attendant.emergency_phone or '')
        ws.cell(row=row, column=10, value=attendant.created_at.strftime('%Y-%m-%d'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendants_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


def export_events_csv(events):
    """Export events to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="events_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Name', 'Type', 'Location', 'Start Date', 'End Date', 
        'Status', 'Total Stations', 'Expected Attendants', 'Assignments Count'
    ])
    
    for event in events:
        writer.writerow([
            event.name,
            event.get_event_type_display(),
            event.location,
            event.start_date.strftime('%Y-%m-%d'),
            event.end_date.strftime('%Y-%m-%d'),
            event.get_status_display(),
            event.total_stations or 0,
            event.expected_attendants or 0,
            event.assignments.count()
        ])
    
    return response


def export_assignments_csv(assignments):
    """Export assignments to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="assignments_{datetime.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Attendant', 'Event', 'Position', 'Shift Start', 'Shift End', 
        'Duration (Hours)', 'Notes', 'Created Date'
    ])
    
    for assignment in assignments:
        duration = ''
        if assignment.shift_start and assignment.shift_end:
            delta = assignment.shift_end - assignment.shift_start
            duration = f"{delta.total_seconds() / 3600:.1f}"
        
        writer.writerow([
            assignment.attendant.get_full_name(),
            assignment.event.name,
            assignment.position,
            assignment.shift_start.strftime('%Y-%m-%d %H:%M') if assignment.shift_start else '',
            assignment.shift_end.strftime('%Y-%m-%d %H:%M') if assignment.shift_end else '',
            duration,
            assignment.notes or '',
            assignment.created_at.strftime('%Y-%m-%d')
        ])
    
    return response


def export_assignments_excel(assignments):
    """Export assignments to Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignments"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Attendant', 'Event', 'Position', 'Shift Start', 'Shift End', 
        'Duration (Hours)', 'Notes', 'Created Date'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Data rows
    for row, assignment in enumerate(assignments, 2):
        duration = ''
        if assignment.shift_start and assignment.shift_end:
            delta = assignment.shift_end - assignment.shift_start
            duration = f"{delta.total_seconds() / 3600:.1f}"
        
        ws.cell(row=row, column=1, value=assignment.attendant.get_full_name())
        ws.cell(row=row, column=2, value=assignment.event.name)
        ws.cell(row=row, column=3, value=assignment.position)
        ws.cell(row=row, column=4, value=assignment.shift_start.strftime('%Y-%m-%d %H:%M') if assignment.shift_start else '')
        ws.cell(row=row, column=5, value=assignment.shift_end.strftime('%Y-%m-%d %H:%M') if assignment.shift_end else '')
        ws.cell(row=row, column=6, value=duration)
        ws.cell(row=row, column=7, value=assignment.notes or '')
        ws.cell(row=row, column=8, value=assignment.created_at.strftime('%Y-%m-%d'))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="assignments_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    
    return response


def export_assignments_pdf(assignments, title="Assignment Report"):
    """Export assignments to PDF format"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="assignments_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))
    
    # Subtitle with date
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        alignment=1,
        spaceAfter=20
    )
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y')}", subtitle_style))
    elements.append(Spacer(1, 12))
    
    # Table data
    data = [['Attendant', 'Event', 'Position', 'Shift Start', 'Shift End', 'Duration']]
    
    for assignment in assignments:
        duration = ''
        if assignment.shift_start and assignment.shift_end:
            delta = assignment.shift_end - assignment.shift_start
            duration = f"{delta.total_seconds() / 3600:.1f}h"
        
        data.append([
            assignment.attendant.get_full_name(),
            assignment.event.name,
            assignment.position,
            assignment.shift_start.strftime('%m/%d %H:%M') if assignment.shift_start else '',
            assignment.shift_end.strftime('%m/%d %H:%M') if assignment.shift_end else '',
            duration
        ])
    
    # Create table
    table = Table(data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1*inch, 1*inch, 0.7*inch])
    
    # Table style
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    return response


def export_event_schedule_pdf(event):
    """Export detailed event schedule to PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="schedule_{event.name.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=20,
        alignment=1
    )
    elements.append(Paragraph(f"{event.name} - Schedule", title_style))
    
    # Event details
    event_info = f"""
    <b>Location:</b> {event.location}<br/>
    <b>Date:</b> {event.start_date.strftime('%B %d, %Y')} - {event.end_date.strftime('%B %d, %Y')}<br/>
    <b>Type:</b> {event.get_event_type_display()}<br/>
    <b>Status:</b> {event.get_status_display()}
    """
    elements.append(Paragraph(event_info, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Assignments table
    assignments = event.assignments.all().order_by('shift_start', 'position')
    
    if assignments:
        data = [['Position', 'Attendant', 'Shift Time', 'Contact', 'Notes']]
        
        for assignment in assignments:
            shift_time = ''
            if assignment.shift_start and assignment.shift_end:
                shift_time = f"{assignment.shift_start.strftime('%H:%M')} - {assignment.shift_end.strftime('%H:%M')}"
            
            data.append([
                assignment.position,
                assignment.attendant.get_full_name(),
                shift_time,
                assignment.attendant.phone or '',
                assignment.notes[:50] + '...' if assignment.notes and len(assignment.notes) > 50 else assignment.notes or ''
            ])
        
        table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1.2*inch, 1*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph("No assignments found for this event.", styles['Normal']))
    
    doc.build(elements)
    return response
