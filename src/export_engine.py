"""
Export Engine for JW Attendant Scheduler

Advanced export capabilities for generating schedules in multiple formats.
Supports PDF, Excel, and CSV exports with customizable templates.
"""

import os
import csv
from datetime import datetime
from typing import List, Dict, Any
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import Event, Assignment, Volunteer


class ExportEngine:
    """Advanced export engine for multiple format support"""
    
    def __init__(self):
        self.export_dir = 'data/exports'
        self.ensure_export_directory()
        
        # PDF styles
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        self.header_style = ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12
        )
    
    def ensure_export_directory(self):
        """Ensure export directory exists"""
        if not os.path.exists(self.export_dir):
            os.makedirs(self.export_dir)
    
    def export_pdf_schedule(self, event: Event, assignments: List[Assignment]) -> str:
        """Export schedule as PDF"""
        filename = f"schedule_{event.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        filepath = os.path.join(self.export_dir, filename)
        
        # Create PDF document
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        
        # Title
        title = Paragraph(f"Attendant Schedule - {event.name}", self.title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Event details
        event_details = [
            f"<b>Event Type:</b> {event.event_type.value}",
            f"<b>Dates:</b> {event.start_date} to {event.end_date}",
            f"<b>Location:</b> {event.location or 'TBD'}",
            f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        
        for detail in event_details:
            story.append(Paragraph(detail, self.styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Group assignments by date and time
        grouped_assignments = self._group_assignments_by_datetime(assignments)
        
        for date_time, date_assignments in grouped_assignments.items():
            # Date/time header
            header = Paragraph(f"<b>{date_time}</b>", self.header_style)
            story.append(header)
            
            # Create table data
            table_data = [['Position', 'Volunteer', 'Experience', 'Contact', 'Notes']]
            
            for assignment in date_assignments:
                volunteer = assignment.volunteer
                table_data.append([
                    assignment.position,
                    f"{volunteer.first_name} {volunteer.last_name}",
                    volunteer.experience_level or 'N/A',
                    volunteer.phone or volunteer.email or 'N/A',
                    assignment.notes or ''
                ])
            
            # Create and style table
            table = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1.5*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(story)
        return filepath
    
    def export_excel_schedule(self, event: Event, assignments: List[Assignment]) -> str:
        """Export schedule as Excel file"""
        filename = f"schedule_{event.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"
        
        # Title and event info
        ws['A1'] = f"Attendant Schedule - {event.name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        ws['A3'] = f"Event Type: {event.event_type.value}"
        ws['A4'] = f"Dates: {event.start_date} to {event.end_date}"
        ws['A5'] = f"Location: {event.location or 'TBD'}"
        ws['A6'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Headers
        headers = ['Date', 'Time', 'Position', 'Volunteer', 'Experience', 'Contact', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = 9
        for assignment in sorted(assignments, key=lambda a: a.shift_start):
            volunteer = assignment.volunteer
            
            ws.cell(row=row, column=1, value=assignment.shift_start.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=f"{assignment.shift_start.strftime('%H:%M')} - {assignment.shift_end.strftime('%H:%M')}")
            ws.cell(row=row, column=3, value=assignment.position)
            ws.cell(row=row, column=4, value=f"{volunteer.first_name} {volunteer.last_name}")
            ws.cell(row=row, column=5, value=volunteer.experience_level or 'N/A')
            ws.cell(row=row, column=6, value=volunteer.phone or volunteer.email or 'N/A')
            ws.cell(row=row, column=7, value=assignment.notes or '')
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=8, max_row=row-1, min_col=1, max_col=7):
            for cell in row:
                cell.border = thin_border
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        self._create_excel_summary(summary_ws, event, assignments)
        
        # Create volunteer sheet
        volunteer_ws = wb.create_sheet("Volunteers")
        self._create_excel_volunteer_sheet(volunteer_ws, assignments)
        
        wb.save(filepath)
        return filepath
    
    def export_csv_schedule(self, event: Event, assignments: List[Assignment]) -> str:
        """Export schedule as CSV file"""
        filename = f"schedule_{event.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(self.export_dir, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Header information
            writer.writerow([f"Attendant Schedule - {event.name}"])
            writer.writerow([f"Event Type: {event.event_type.value}"])
            writer.writerow([f"Dates: {event.start_date} to {event.end_date}"])
            writer.writerow([f"Location: {event.location or 'TBD'}"])
            writer.writerow([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])  # Empty row
            
            # Column headers
            writer.writerow(['Date', 'Start Time', 'End Time', 'Position', 'Volunteer Name', 
                           'Experience Level', 'Phone', 'Email', 'Congregation', 'Notes'])
            
            # Data rows
            for assignment in sorted(assignments, key=lambda a: a.shift_start):
                volunteer = assignment.volunteer
                writer.writerow([
                    assignment.shift_start.strftime('%Y-%m-%d'),
                    assignment.shift_start.strftime('%H:%M'),
                    assignment.shift_end.strftime('%H:%M'),
                    assignment.position,
                    f"{volunteer.first_name} {volunteer.last_name}",
                    volunteer.experience_level or 'N/A',
                    volunteer.phone or '',
                    volunteer.email or '',
                    volunteer.congregation or '',
                    assignment.notes or ''
                ])
        
        return filepath
    
    def export_individual_schedules(self, event: Event, assignments: List[Assignment]) -> List[str]:
        """Export individual schedules for each volunteer"""
        filepaths = []
        
        # Group assignments by volunteer
        volunteer_assignments = {}
        for assignment in assignments:
            volunteer_id = assignment.volunteer_id
            if volunteer_id not in volunteer_assignments:
                volunteer_assignments[volunteer_id] = []
            volunteer_assignments[volunteer_id].append(assignment)
        
        # Create individual schedules
        for volunteer_id, vol_assignments in volunteer_assignments.items():
            volunteer = vol_assignments[0].volunteer
            filename = f"individual_schedule_{volunteer.first_name}_{volunteer.last_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(self.export_dir, filename)
            
            # Create PDF document
            doc = SimpleDocTemplate(filepath, pagesize=letter)
            story = []
            
            # Title
            title = Paragraph(f"Personal Schedule - {volunteer.first_name} {volunteer.last_name}", self.title_style)
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Event and volunteer details
            details = [
                f"<b>Event:</b> {event.name}",
                f"<b>Event Type:</b> {event.event_type.value}",
                f"<b>Dates:</b> {event.start_date} to {event.end_date}",
                f"<b>Location:</b> {event.location or 'TBD'}",
                f"<b>Volunteer:</b> {volunteer.first_name} {volunteer.last_name}",
                f"<b>Experience Level:</b> {volunteer.experience_level or 'N/A'}",
                f"<b>Congregation:</b> {volunteer.congregation or 'N/A'}"
            ]
            
            for detail in details:
                story.append(Paragraph(detail, self.styles['Normal']))
            
            story.append(Spacer(1, 20))
            
            # Assignments table
            table_data = [['Date', 'Time', 'Position', 'Notes']]
            
            for assignment in sorted(vol_assignments, key=lambda a: a.shift_start):
                table_data.append([
                    assignment.shift_start.strftime('%Y-%m-%d'),
                    f"{assignment.shift_start.strftime('%H:%M')} - {assignment.shift_end.strftime('%H:%M')}",
                    assignment.position,
                    assignment.notes or ''
                ])
            
            table = Table(table_data, colWidths=[1.5*inch, 2*inch, 2*inch, 2.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10)
            ]))
            
            story.append(table)
            
            # Build PDF
            doc.build(story)
            filepaths.append(filepath)
        
        return filepaths
    
    def _group_assignments_by_datetime(self, assignments: List[Assignment]) -> Dict[str, List[Assignment]]:
        """Group assignments by date and time"""
        grouped = {}
        
        for assignment in assignments:
            date_str = assignment.shift_start.strftime('%A, %B %d, %Y')
            time_str = f"{assignment.shift_start.strftime('%H:%M')} - {assignment.shift_end.strftime('%H:%M')}"
            key = f"{date_str} ({time_str})"
            
            if key not in grouped:
                grouped[key] = []
            grouped[key].append(assignment)
        
        # Sort by datetime
        return dict(sorted(grouped.items(), key=lambda x: x[1][0].shift_start))
    
    def _create_excel_summary(self, ws, event: Event, assignments: List[Assignment]):
        """Create summary sheet in Excel workbook"""
        ws['A1'] = "Schedule Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Statistics
        total_assignments = len(assignments)
        unique_volunteers = len(set(a.volunteer_id for a in assignments))
        positions = list(set(a.position for a in assignments))
        
        ws['A3'] = "Statistics:"
        ws['A3'].font = Font(bold=True)
        ws['A4'] = f"Total Assignments: {total_assignments}"
        ws['A5'] = f"Unique Volunteers: {unique_volunteers}"
        ws['A6'] = f"Positions Used: {len(positions)}"
        
        # Position breakdown
        ws['A8'] = "Position Breakdown:"
        ws['A8'].font = Font(bold=True)
        
        position_counts = {}
        for assignment in assignments:
            position_counts[assignment.position] = position_counts.get(assignment.position, 0) + 1
        
        row = 9
        for position, count in sorted(position_counts.items()):
            ws[f'A{row}'] = f"{position}: {count}"
            row += 1
    
    def _create_excel_volunteer_sheet(self, ws, assignments: List[Assignment]):
        """Create volunteer summary sheet in Excel workbook"""
        ws['A1'] = "Volunteer Summary"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Headers
        headers = ['Volunteer', 'Total Assignments', 'Positions', 'Total Hours', 'Experience']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Volunteer data
        volunteer_data = {}
        for assignment in assignments:
            volunteer = assignment.volunteer
            vol_key = f"{volunteer.first_name} {volunteer.last_name}"
            
            if vol_key not in volunteer_data:
                volunteer_data[vol_key] = {
                    'assignments': 0,
                    'positions': set(),
                    'hours': 0,
                    'experience': volunteer.experience_level or 'N/A'
                }
            
            volunteer_data[vol_key]['assignments'] += 1
            volunteer_data[vol_key]['positions'].add(assignment.position)
            hours = (assignment.shift_end - assignment.shift_start).total_seconds() / 3600
            volunteer_data[vol_key]['hours'] += hours
        
        row = 4
        for volunteer, data in sorted(volunteer_data.items()):
            ws.cell(row=row, column=1, value=volunteer)
            ws.cell(row=row, column=2, value=data['assignments'])
            ws.cell(row=row, column=3, value=', '.join(sorted(data['positions'])))
            ws.cell(row=row, column=4, value=f"{data['hours']:.1f}")
            ws.cell(row=row, column=5, value=data['experience'])
            row += 1
